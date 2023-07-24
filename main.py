from PIL import Image, ImageDraw, ImageFont
import openpyxl
import concurrent.futures
import time

start = time.perf_counter()

def iterate_first_column(file_path, sheet_name):
    numbers = []
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Select the sheet by name
    sheet = wb[sheet_name]

    # Get the maximum row count (number of rows with data)
    max_row = sheet.max_row

    # Iterate through the cells in the first column
    for row in range(1, max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value is not None:
            numbers.append(cell_value)

    # Close the workbook
    wb.close()
    return numbers

# Usage example
file_path = "base/lista.xlsx"
sheet_name = "Sheet1"

labels = iterate_first_column(file_path, sheet_name)
# print(labels)
# If there are not 888 labes, which is the number of labes wanted, exit the program
print("Numero de etiquetas: " + str(len(labels)))
if len(labels) != 888:
    print("Error: el numero de etiquetas no es 888. Comprueba que el archivo lista.xlsx tiene las"
          " 888 etiquetas.")
    exit()

# Get the first 10 labels to do a small test
# labels = labels[0:10]

num_threads = 8
image_path = "base/base.png"
def poner_etiqueta(label):
    # Get the base image
    image = Image.open(image_path)
    # Get a drawing context
    draw = ImageDraw.Draw(image)

    # Draw text
    text = str(label)
    font_path = "base/Glacial_Indifference.otf"  # Replace with the actual path to your .ttf font file
    font_size = 130
    font = ImageFont.truetype(font_path, font_size)
    text_color = (59, 196, 160)  # RGB color tuple for the color of the delegation

    # Get the width and height of the image
    image_width, image_height = image.size

    # Get size of text, so it is centered
    text_bbox = draw.textbbox((0, 0), text, font=font)
    text_x = (image_width - text_bbox[2]) // 2
    text_y = 320

    # Draw the text and save it to the output path
    draw.text((text_x, text_y), text, font=font, fill=text_color)
    output_path = "output/" + str(label) + "_pegatina.png"
    image.save(output_path)
    print("Etiqueta " + str(label) + " generada.")


# Do the work in parallel using a thread pool
with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
    futures = {executor.submit(poner_etiqueta, item): item for item in labels}

    # Wait for all tasks to complete
    concurrent.futures.wait(futures)

finish = time.perf_counter()
print(f'Finished in {round(finish-start, 2)} second(s)')