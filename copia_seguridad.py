from PIL import Image, ImageDraw, ImageFont, ImageColor
import openpyxl
import time

start = time.perf_counter()

def iterate_first_column(file_path, sheet_name):
    numbers = []
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Select the sheet by name
    sheet = wb[sheet_name]

    # Get the maximum row count (number of rows with data)
    max_row = sheet.max_row

    # Iterate through the cells in the first column (column A)
    for row in range(1, max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value is not None:
            #print(cell_value)
            numbers.append(cell_value)

    # Close the workbook
    wb.close()
    return numbers

# Usage example
file_path = "base/lista.xlsx"
sheet_name = "Sheet1"

labels = iterate_first_column(file_path, sheet_name)
print(labels)
print("Numero de etiquetas: " + str(len(labels)))

image_path = "base/base.png"

global counter
counter = 0
for label in labels:

    image = Image.open(image_path)
    #Get a drawing context
    draw = ImageDraw.Draw(image)

    # Draw text
    text = str(label)
    font_path = "base/Exo-Regular.ttf"  # Replace with the actual path to your .ttf font file
    font_size = 130
    font = ImageFont.truetype(font_path, font_size)
    text_color = (59, 196, 160)  # RGB color tuple for the color of the delegation

    # Get the width and height of the image
    image_width, image_height = image.size

    # Get size of text
    text_width, text_height = draw.textsize(text, font=font)

    text_x = (image_width - text_width) // 2
    text_y = 320

    draw.text((text_x, text_y), text, font=font, fill=text_color)
    output_path = "salida_no/" + str(label) + "_pegatina.png"
    image.save(output_path)
    counter += 1
    print("Etiqueta " + str(label) + " generada. Llevo "+ str(counter) + " de " + str(len(labels)))

stop = time.perf_counter()
print(f"Tiempo de ejecucion: {stop - start:0.4f} segundos")